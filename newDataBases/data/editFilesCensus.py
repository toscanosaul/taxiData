import csv
from openpyxl import load_workbook
import re
from openpyxl import load_workbook
from geopy.geocoders import Nominatim


highSchoolFile="DOE_High_School_Directory_2014-2015.csv"
hFile="highSchool.csv"

econFile="Demographics_20and_20profiles_20at_20the_20Neighborhood_20Tabulation_20Area_20_NTA__20level/acs_select_econ_08to12_ntas.xlsx"
eFile="econFile.csv"


def highSchool():
    geolocator = Nominatim()
    with open(highSchoolFile,"rb") as school:
        f=csv.reader(school)
        with open(hFile,"wb") as result:
            wtr=csv.writer(result)
            row=0
            for r in f:
                if row>0:
                    temp=r[12]+","+r[13]+","+r[14]
                    if r[17]=="N/A":
                        continue
                    location = geolocator.geocode(temp)
                    if location is None:
                        s=r[12].split(" ")
                        if s[-1]=="Street" or s[-1]=="street":
                            if s[len(s)-2].isdigit():
                                tmp=s[len(s)-2]
                                nTmp=len(tmp)
                                if nTmp!=2 or tmp[0]!="1":
                                    if tmp[nTmp-1]=="1":
                                        s[len(s)-2]+="st"
                                    elif tmp[nTmp-1]=="2":
                                        s[len(s)-2]+="nd"
                                    elif tmp[nTmp-1]=="3":
                                        s[len(s)-2]+="rd"
                                    else:
                                        s[len(s)-2]+="th"
                                else:
                                    s[len(s)-2]+="th"
                        temp=" ".join(s)+","+r[13]+","+r[14]
                        location = geolocator.geocode(temp)
                    if location is not None:
                        wtr.writerow((location.latitude,location.longitude,r[17]))
                    else:
                        print "fail"
                        print temp
                else:
                    wtr.writerow(("latitude","longitude","total_students"))
                    row=1
            
            

cols=[]
wb = load_workbook(econFile,True)
ws=wb['sheet2']


i=0
for row in ws.iter_rows():
    if i<5:
        i+=1
    else:
        number=[]
        col=0
        for cell in row:
            if col<1:
                col+=1
            else:
                if cell.internal_value is not None and col%4==1:
                    number.append(cell.internal_value[0:4])
                col+=1
        break

start,end=62,72

cols=[[] for i in range(10)]
for index,row in enumerate(ws.iter_rows()):
    if start<=index<end:
        col=0
        for cell in row:
            if col<1:
                col+=1
            else:
                if col%4==1:
                    temp=cell.internal_value
                    if temp is not None:
                        cols[index-start].append(cell.internal_value)
                col+=1
    if index>=end:
        break

salary=["<10,000","10,000-14,999","15,000-24,999",
        "25,000-34,999","35,000-49,999","50,000-74,999",
        "75,000-99,999","100,000-149,999","150,000-199,999",
        ">200,000"]

with open(eFile,"wb") as f:
    out=csv.writer(f)
    out.writerow(["ntacode"]+salary)
    #out.writerow(["Salary"]+number)
    for i in range(len(number)):
        temp=[]
        for j in range(10):
            temp+=[cols[j][i]]
        out.writerow([number[i]]+temp)




